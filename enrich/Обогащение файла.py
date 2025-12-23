import json
import re
import numpy as np
import pandas as pd
import os
import glob
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from enrich.main2 import enrich_dataframe, load_json
from functions import get_unique_filepath

"""
Данный скрипт заполняет эксель файл данными из базы неизменяемых данных, изменяемых данных и квартирографии
"""

# Загружаем Excel
file_path = r"C:\Users\m.olshanskiy\PycharmProjects\ndv_parsing\enrich\output.xlsx"

df = pd.read_excel(file_path)

projects_dict = load_json(
        r'C:\Users\m.olshanskiy\PycharmProjects\ndv_parsing\!haracteristik_dictionary\projects.json'
)
corpus_dict = load_json(
    r'C:\Users\m.olshanskiy\PycharmProjects\ndv_parsing\!changing_haracteristik_dictionary\projects.json'
)
area_dict = load_json(r'C:\Users\m.olshanskiy\PycharmProjects\ndv_parsing\area_dictionary\output.json'
)

df = enrich_dataframe(
    df,
    projects_dict=projects_dict,
    corpus_dict=corpus_dict,
    area_dict=area_dict
)  # если указать =None, то изменений не будет

df.to_excel(file_path, index=False)

# открываем файл и применяем выравнивание
wb = load_workbook(file_path)
ws = wb.active

center_alignment = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = center_alignment

# Автоподбор ширины колонок
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)  # Получаем буквенное обозначение колонки

    for cell in col:
        try:
            if cell.value:
                # Учитываем длину текста в ячейке
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        except:
            pass

    # Устанавливаем ширину колонки (добавляем 2 для отступов)
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(file_path)
print(f"✅ Данные сохранены в файл: {file_path}")


















